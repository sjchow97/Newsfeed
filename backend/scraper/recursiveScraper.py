import requests
from bs4 import BeautifulSoup
from urlparse import urljoin, urlparse
import urllib2
import HTMLParser
import re
import time
import openpyxl
import os
import sys

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

# Set the Django settings module environment variable
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'newsfeed.settings')

# Import Django and set up the environment
import django
django.setup()
from rss.models import RssSource

import sqlite3
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Join the path to the SQLite database file inside the 'backend' directory
db_path = os.path.join(BASE_DIR, 'db.sqlite3')

conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Clear the RssSource table before inserting new data to avoid duplicates when reseeding
cursor.execute("DELETE FROM rss_rsssource WHERE source_id >= 0;")

checked_urls = set()
found_rss_feeds = set()

def find_rss_feed(url, municipality, province):
    global checked_urls
    global found_rss_feeds
    try:
        if url in checked_urls:
            print('URL already checked:', url)
            return
        checked_urls.add(url)

        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            # Look for <link> and <a> tags
            rss_links = soup.find_all(['link', 'a'])
            if rss_links:
                for link in rss_links:
                    rss_url = link.get('href')
                    # Check if the URL contains "rss" or "feed/" (important to check for feed/ otherwise common URLs like /feedback will be included as well)
                    # Flag the URL as a potential RSS feed if it contains the keywords
                    if rss_url and ('rss' in rss_url.lower() or 'feed/' in rss_url.lower()):
                        full_rss_url = urljoin(url, rss_url)
                        if full_rss_url not in found_rss_feeds:
                            
                            print('Potential RSS Feed on the page:  {}: {}'.format(url, full_rss_url))

                            try:
                                response = urllib2.urlopen(full_rss_url)

                                if response.getcode() == 200:
                                    html = response.read()
                                    soup = BeautifulSoup(html, 'html.parser')

                                    # find whether or not the flagged content is a RSS feed, which will contain the <rss> or <feed> tag
                                    if "<rss" in html or "<feed" in html:
                                       
                                        print('RSS feed found on {}: {}'.format(url, full_rss_url))
                                        found_rss_feeds.add(full_rss_url)

                                        # These fields are related to the RssSource model
                                        print("BASE URL: " + full_rss_url)
                                        print("NAME: " + soup.title.string)
                                        print("LOCATION: " + municipality + ", " + province)

                                        items = soup.find_all('item')
                                        print "NUMBER OF NEWS ITEMS: " + str(len(items))

                                        # if there are items in the RSS feed, create a new RssSource object and save it to the database
                                        if len(items) > 0:
                                            
                                            # create a new RssSource object
                                            rss_source = RssSource()
                                            rss_source.source_name = soup.title.string
                                            rss_source.url = full_rss_url
                                            rss_source.location = municipality + ", " + province

                                            
                                            cursor.execute("INSERT INTO rss_rsssource (source_name, url, location) VALUES (?, ?, ?);", (rss_source.source_name, rss_source.url, rss_source.location))
                                            conn.commit()

                                        # parse each <item> tag for fields that will be displayed on the news feed
                                        for item in items:
                                            title = item.find('title')
                                            description = item.find('description').get_text()
                                            link = item.find('link')
                                            pub_date = item.find('pubdate')

                                            # unescape the title to remove any html entities and encode it to utf-8
                                            unescaped_title = (HTMLParser.HTMLParser().unescape(title.get_text())).encode('utf-8')
                                            print "Title: " + unescaped_title
                                            
                                            # do not include items that do not have a pubdate onto the news feed
                                            if pub_date == None:
                                                print "No publication date found for this news item, skipping to next item"
                                                continue
                                            else:
                                                print "Publication date: " + pub_date.get_text()

                                            # unescape the description to remove any html entities and encode it to utf-8
                                            unescaped_description = (HTMLParser.HTMLParser().unescape(description)).encode('utf-8')
                                            print "Description: " + unescaped_description
                                    else:
                                        
                                        # if the flagged url is not an rss feed, perform a recursive search into the flagged page's urls as some cases will have multiple feeds listed on the /rss page such as coquitlam https://www.coquitlam.ca/rss.aspx
                                        find_rss_feed(full_rss_url, municipality, province)
                                        
                            except urllib2.HTTPError as e:
                            # If an HTTP error occurs, print the status code and error message
                                print("HTTP Error:", e.code, e.reason)
            else:
                print('No RSS feed found on:', url)
        else:
            print('Failed to fetch {}: Status code {}'.format(url, response.status_code))
    except Exception as e:
        print('Error occurred while fetching {}: {}'.format(url, e))


def crawl_website(start_url, municipality, province):
    visited_urls = set()
    queue = [start_url]

    print 'Crawling website for: ' + municipality + ', ' + province

    # For demo purposes, each site will be crawled for 30 seconds, set timeout to larger value to increase number of pages crawled at expense of runtime
    start_time = time.time()
    timeout = 25

    while queue:

        # Check if the time elapsed for crawling this site is greater than the timeout value
        if time.time() - start_time > timeout:
            print('Timeout reached. Stopping the crawl.')
            break


        url = queue.pop(0)
        if url not in visited_urls:
            find_rss_feed(url, municipality, province)
            visited_urls.add(url)
            try:
                response = requests.get(url)
                if response.status_code == 200:
                    soup = BeautifulSoup(response.content, 'html.parser')
                    # Find all <a> tags and enqueue their URLs for further crawling
                    for link in soup.find_all('a', href=True):
                        absolute_url = urljoin(url, link['href'])
                        parsed_url = urlparse(absolute_url)
                        base_path = urlparse(start_url).path

                        # Ensures that the script only check the URLs within the same domain
                        if absolute_url.startswith(start_url):
                            queue.append(absolute_url)
            except Exception as e:
                print 'Error occurred while crawling {}: {}'.format(url, e)


# Open the Excel file containing the list of municipalities and pass the URL of the official website to the crawl_website function
municipalities = openpyxl.load_workbook('municipalities.xlsx')
for sheet in municipalities.worksheets:
    
    # Extract the province from the sheet name
    province = sheet.title
    
    # Skip the header row by starting from row 2
    for row in sheet.iter_rows(min_row=2):
    
        # Get the municipality and URL
        municipality = row[0].value
        url = row[0].hyperlink.target
        print ("URL: " + url)
        print ("Municipality: " + municipality)
    
        try:
            response = urllib2.urlopen(url)
    
            if response.getcode() == 200:
                html = response.read()
                soup = BeautifulSoup(html, 'html.parser')
    
                # Find the <a> tag within the <span> tags with class "official-website" and get the href attribute
                link = soup.find('span', class_='official-website').find('a').get('href')
                    
                crawl_website(link, municipality, province)
    
        except urllib2.HTTPError as e:
        # If an HTTP error occurs, print the status code and error message
            print("HTTP Error:", e.code, e.reason)



# Open the Excel file containing the list of media outlets and pass the URL of the official website to the crawl_website function
media_outlets = openpyxl.load_workbook('media_outlets.xlsx')
for sheet in media_outlets.worksheets:

    # Extract the province from the sheet name
    province = sheet.title

    # Skip the header row by starting from row 2
    for row in sheet.iter_rows(min_row=2):

        # Get the URL and municipality
        url = row[0].hyperlink.target
        municipality = row[1].value

        try:
            crawl_website(url, municipality, province)
        except urllib2.HTTPError as e:
            # If an HTTP error occurs, print the status code and error message
            print("HTTP Error:", e.code, e.reason)



cursor.close()
conn.close()